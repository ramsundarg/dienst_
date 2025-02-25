import calendar
import datetime
import glob
from pathlib import Path
import re
import base64
import http.client
import json
import urllib
from time import sleep
import os
import pandas as pd
import requests
from ics import Calendar, Event
import openpyxl
from openpyxl.styles import PatternFill

publishIcs = True
c = Calendar()

Employee = 'TRG'
convertFiles = True

def convert_files():
    for file_name in list(glob.glob('original_data/*.pdf')):
        dest_file = Path(file_name).stem
        with open(file_name, 'rb') as fp:
            payload = fp.read()
            convert_file(file_name, payload)
        print("All files converted, now computing the dienst")

def convert_file(file_name, file_content):
    import http
    conn = http.client.HTTPSConnection("pdf-services-ew1.adobe.io")

    payload = "client_id=540f46c8507b47998d1c238878658535&client_secret=p8e-eNckbaGuPkDLT6iUEYSsye8ANTL-3ljD"

    headers = {
        'Content-Type': "application/x-www-form-urlencoded",
        'User-Agent': "insomnia/8.5.1"
    }

    conn.request("POST", "/token", payload, headers)

    res = conn.getresponse()
    data = res.read()

    data_decoded = data.decode("utf-8")
    data_json = json.loads(data_decoded)
    access_token = data_json['access_token']

    conn = http.client.HTTPSConnection("pdf-services.adobe.io")

    payload = "{\n\t\"mediaType\" :  \"application/pdf\"\n}"

    headers = {
        'Content-Type': "application/json",
        'User-Agent': "insomnia/8.5.1",
        'Authorization': f"Bearer {access_token}",
        'x-api-key': "540f46c8507b47998d1c238878658535"
    }

    conn.request("POST", "/assets", payload, headers)

    res = conn.getresponse()
    data = res.read()
    data_decoded = data.decode("utf-8")
    data_json = json.loads(data_decoded)
    uploadUri = data_json['uploadUri']
    assetId = data_json['assetID']

    import http.client

    share_host = 'dcplatformstorageservice-prod-us-east-1.s3-accelerate.amazonaws.com'
    conn = http.client.HTTPSConnection(share_host)
    share_host = f"https://{share_host}"

    dest_file = Path(file_name).stem
    payload = file_content

    headers = {
        'Content-Type': "application/pdf",
        'User-Agent': "insomnia/8.5.1"
    }

    conn.request("PUT", uploadUri.removeprefix(share_host), payload, headers)
    res = conn.getresponse()

    conn = http.client.HTTPSConnection("pdf-services-ue1.adobe.io")

    payload = {"assetID": assetId, "targetFormat": "xlsx", "ocrLang": "de-DE"}
    payload = json.dumps(payload)

    headers = {
        'Content-Type': "application/json",
        'User-Agent': "insomnia/8.5.1",
        'Authorization': f"Bearer {access_token}",
        'x-api-key': "540f46c8507b47998d1c238878658535"
    }

    conn.request("POST", "/operation/exportpdf", payload, headers)

    res = conn.getresponse()
    request_id = res.headers['x-request-id']

    conn = http.client.HTTPSConnection("pdf-services-ue1.adobe.io")

    payload = ""

    headers = {
        'User-Agent': "insomnia/8.5.1",
        'Authorization': f"Bearer {access_token}",
        'x-api-key': "540f46c8507b47998d1c238878658535"
    }

    while True:
        conn.request("GET", f"/operation/exportpdf/{request_id}/status?=", payload, headers)

        res = conn.getresponse()
        data = res.read()
        data_decoded = data.decode("utf-8")
        data_json = json.loads(data_decoded)
        print(data_json)
        downloadUri = data_json.get('asset', '')
        if downloadUri == '':
            print(f"Not converted file {file_name}, waiting for 3 seconds")
            sleep(3)
        else:
            urllib.request.urlretrieve(downloadUri['downloadUri'], f'{dest_file}.xlsx')
            return f'{dest_file}.xlsx'

def get_df(files, name, year, month):
    rows = []
    dates = set()
    for file_name in files:
        xl = pd.ExcelFile(file_name)
        res = len(xl.sheet_names)
        work_type = Path(file_name).stem
        regex = re.compile('[^a-zA-Z]')
        work_type = regex.sub('', work_type)

        work_type_hours = {}
        work_type_desc = {}
        print(f"Reading sheet {file_name}")
        for sheet_name in xl.sheet_names:
            a = pd.read_excel(file_name, sheet_name=sheet_name)
            a1 = list(a.values.flatten())
            a2 = [str(s) for s in a1 if "=" in str(s)]
            for hours in a2:
                for hour in hours.splitlines():
                    result = re.search(r"(.*) =", hour)
                    # regular expression pattern
                    pattern = r"[0-9]+[:]?[0-9]*-[0-9]+[:]?[0-9]*"
                    work_type_hours[result.group(1)] = {}
                    # find all matches
                    matches = re.findall(pattern, hour)

                    # get the first match
                    first_match = matches[0] if matches else None

                    if first_match:
                        # split the string on '-'
                        before_dash, after_dash = first_match.split('-', 1)

                        # if before_dash is of the format number:number
                        if ':' in before_dash:
                            # take the number before the : as hour and number after colon as minute
                            hour_before, minute_before = map(int, before_dash.split(':'))
                        else:
                            # use all the numbers as hour
                            hour_before = int(before_dash)
                            minute_before = 0

                        # construct a datetime
                        datetime_before = datetime.time(hour=hour_before, minute=minute_before)

                        # if after_dash is of the format number:number
                        if ':' in after_dash:
                            # take the number before the : as hour and number after colon as minute
                            hour_after, minute_after = map(int, after_dash.split(':'))
                        else:
                            # use all the numbers as hour
                            hour_after = int(after_dash)
                            minute_after = 0

                        # construct a datetime
                        datetime_after = datetime.time(hour=hour_after, minute=minute_after)
                        work_type_hours[result.group(1)]['start'] = datetime_before.strftime("%H:%M")
                        work_type_hours[result.group(1)]['end'] = datetime_after.strftime("%H:%M")

                    work_type_desc[result.group(1)] = hour.split("=", 1)[1]

        a = pd.read_excel(file_name, header=None)
        skip_rows = a[a[0] == 'Datum'].index[0]
        for sheet_name in xl.sheet_names:
            a = pd.read_excel(file_name, skiprows=skip_rows, header=0, sheet_name=sheet_name)
            emp = a[a[a.columns[0]].str.contains(name, na=False)]
            if emp.empty:
                continue
            for i in range(1, 32):
                if i in dates:
                    continue
                try:
                    date1 = datetime.datetime(year, month, i, 0, 0)
                except ValueError:
                    continue
                value = ""
                Uhr = ""

                if emp.get(i, None).all():
                    value = emp[i].values[0]
                    if value == 'D':
                        continue
                    elif (emp[i].isna().values[0]):
                        rows.append({'date': date1, 'day': calendar.day_name[date1.weekday()], "start": "", "end": "",
                                     "work_type_code": "Free",
                                     'work_type_desc': "", "work_type": ""
                                     })
                        e = Event()
                        e.name = "Free"
                        e.begin = date1.strftime("%Y-%m-%d") + "T00:00:00"
                        c.events.add(e)

                        # [<Event 'My cool event' begin:2014-01-01 00:00:00 end:2014-01-01 00:00:01>]
                        with open('my.ics', 'w') as my_file:
                            my_file.writelines(c.serialize_iter())
                    else:
                        e = Event()
                        e.name = f"{Employee}-{work_type}-{value}"
                        start_time = work_type_hours.get(value, {}).get('start', "")
                        end_time = work_type_hours.get(value, {}).get('end', "")
                        if start_time == "" or end_time == "":
                            e.begin = date1.strftime("%Y-%m-%d") + f"T00:00"
                            e.end = date1.strftime("%Y-%m-%d") + f"T23:59"
                        else:
                            begin = date1.strftime("%Y-%m-%d") + f"T{start_time}:00"
                            end = date1.strftime("%Y-%m-%d") + f"T{end_time}:00"
                            begin_time = datetime.datetime.strptime(begin, "%Y-%m-%dT%H:%M:%S")
                            end_time = datetime.datetime.strptime(end, "%Y-%m-%dT%H:%M:%S")
                            if begin_time > end_time:
                                end_time = end_time + datetime.timedelta(days=1)
                            e.begin = begin_time
                            e.end = end_time
                        c.events.add(e)
                        rows.append(
                            {'date': date1, 'day': calendar.day_name[date1.weekday()],
                             "start": work_type_hours.get(value, {}).get('start', ""),
                             "end": work_type_hours.get(value, {}).get('end', ""), "work_type_code": value,
                             'work_type_desc': work_type_desc.get(value, value),
                             "work_type": work_type})
                    dates.add(i)
    df_processed = pd.DataFrame.from_records(rows)
    df_processed = df_processed.sort_values(by='date').drop_duplicates('date', keep='first')
    return df_processed

def adjust_column_width(ws):
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

# Adjust column width for better readability
def apply_styling_to_excel(excel_path):
    # Load the workbook and sheet
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # Define the fill for 'Free' cells
    dark_green_fill = PatternFill(start_color="006400", end_color="006400", fill_type="solid")
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    # Apply the fill to cells with 'Free' in the 'work_type_code' column
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value == 'Free':
                cell.fill = dark_green_fill

    # Adjust column width
    adjust_column_width(ws)

    # Save the workbook
    wb.save(excel_path)

if __name__ == "__main__":
    convertFiles = False
    name = 'TRG'  # Replace with actual employee name
    year = 2025
    month = 3

    # Example usage
    if convertFiles:
        convert_files()
    
    # Assuming you have some files to process
    files = list(glob.glob("data/*.xlsx"))

    df = get_df(files, name, year, month)
    df['date'] = df['date'].dt.strftime('%Y-%m-%d')
    df.to_excel('processed_data.xlsx', index=False)
    apply_styling_to_excel('processed_data.xlsx')

    